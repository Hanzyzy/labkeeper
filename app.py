import sys
import os

# Fix: Ensure this project's venv site-packages comes FIRST in sys.path
_this_dir = os.path.dirname(os.path.abspath(__file__))
_venv_site = os.path.join(_this_dir, 'venv', 'Lib', 'site-packages')
if _venv_site in sys.path:
    sys.path.remove(_venv_site)
sys.path.insert(0, _venv_site)

"""LabKeeper — Main Flask application
Multi-School Lab Equipment Borrowing System with QR Codes.
- Dual entry portals on landing page: Student Portal & School/Admin Portal
- Student login with "Asal Sekolah" dropdown selection
- Admin manages equipment, students, and borrowings strictly isolated per school
"""
import os
import io
from datetime import datetime, timedelta, timezone
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from dotenv import load_dotenv

WIB = timezone(timedelta(hours=7))


def utcnow() -> datetime:
    """Returns naive datetime in Asia/Jakarta timezone (WIB, UTC+7)."""
    return datetime.now(WIB).replace(tzinfo=None)

from models import db, init_db, Student, Admin, Tool, Borrowing, Config, School
from datetime_utils import get_config
from auth import current_student, current_admin, student_required, admin_required
from qr_utils import generate_qr_for_tool, qr_url_for_tool


def create_app() -> Flask:
    load_dotenv()
    app = Flask(__name__)
    app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "labkeeper-dev-secret-change-me")
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///labkeeper.db"
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)  # session persist 30 hari

    init_db(app)

    # ============= SESSION PERSISTENCE =============
    @app.before_request
    def make_session_permanent():
        if session.get("student_id") or session.get("admin_id"):
            session.permanent = True

    # ============= CONTEXT PROCESSORS =============
    @app.context_processor
    def inject_globals():
        import datetime
        import random
        timestamp = int(datetime.datetime.now().timestamp())
        c_student = current_student()
        c_admin = current_admin()
        schools = School.query.filter_by(is_active=True).order_by(School.name).all()
        return {
            "current_student": c_student,
            "current_admin": c_admin,
            "schools": schools,
            "get_config": get_config,
            "now": utcnow,
            "timestamp": timestamp,
            "random_int": random.randint(10000, 99999),
        }

    # ============= TEMPLATE FILTERS & ERROR HANDLERS =============
    @app.template_filter("timesince")
    def timesince_filter(dt):
        if not dt:
            return ""
        now = utcnow()
        diff = (now - dt).total_seconds()
        return _humanize_seconds(int(diff))

    @app.errorhandler(500)
    def internal_server_error(e):
        db.session.rollback()
        import traceback
        traceback.print_exc()
        flash("Terjadi kesalahan internal pada server. Silakan coba beberapa saat lagi.", "error")
        return redirect(url_for("index"))

    @app.errorhandler(404)
    def page_not_found(e):
        return redirect(url_for("index"))

    # ============= PUBLIC ROUTES (no login) =============
    @app.route("/")
    def index():
        c_student = current_student()
        c_admin = current_admin()
        if c_student:
            return redirect(url_for("student_dashboard"))
        if c_admin:
            return redirect(url_for("admin_dashboard"))

        schools = School.query.filter_by(is_active=True).order_by(School.name).all()
        selected_school_id = schools[0].id if schools else 1
        selected_school = School.query.get(selected_school_id) or (schools[0] if schools else None)

        tools = Tool.query.filter_by(is_active=True, school_id=selected_school_id).order_by(Tool.code).all()
        total_alat = len(tools)
        tersedia = sum(1 for t in tools if t.is_available())
        dipinjam = sum(1 for t in tools if not t.is_available())
        categories = sorted(list(set(t.category for t in tools if t.category)))
        
        return render_template(
            "index.html", 
            tools=tools, 
            total_alat=total_alat, 
            tersedia=tersedia, 
            dipinjam=dipinjam, 
            categories=categories, 
            schools=schools,
            selected_school_id=selected_school_id,
            selected_school=selected_school,
            base_url=get_config().base_url
        )

    @app.route("/tool/<code>")
    def tool_detail(code):
        tool = Tool.query.filter_by(code=code, is_active=True).first()
        if not tool:
            flash(f"Alat dengan kode '{code}' tidak ditemukan di sistem LabKeeper.", "error")
            return redirect(url_for("scan"))
        current = tool.current_borrowing()
        student = current_student()
        history = Borrowing.query.filter_by(tool_id=tool.id).order_by(Borrowing.borrow_date.desc()).limit(20).all()
        return render_template("tool_detail.html", tool=tool, current=current, student=student, borrowing_history=history, base_url=get_config().base_url)

    @app.route("/scan")
    def scan():
        """In-browser QR scanner page for mobile devices."""
        code = request.args.get("code", "").strip().upper()
        tool = None
        if code:
            tool = Tool.query.filter_by(code=code, is_active=True).first()
        return render_template("scan.html", tool=tool)

    # Unified login entry point
    @app.route("/login")
    def login():
        if current_admin():
            return redirect(url_for("admin_dashboard"))
        if current_student():
            return redirect(url_for("student_dashboard"))
        return redirect(url_for("student_login"))

    # ============= STUDENT AUTH + ACTIONS =============
    @app.route("/student/login", methods=["GET", "POST"])
    def student_login():
        if current_admin():
            return redirect(url_for("admin_dashboard"))
        if current_student():
            return redirect(url_for("student_dashboard"))
        
        schools = School.query.filter_by(is_active=True).order_by(School.name).all()
        nxt = request.args.get("next") or url_for("index")
        
        if request.method == "POST":
            school_id = request.form.get("school_id")
            nis = request.form.get("nis", "").strip()
            password = request.form.get("password", "")
            
            if not school_id or not nis or not password:
                flash("Pilih asal sekolah, NIS, dan password wajib diisi.", "warning")
                return render_template("login.html", schools=schools, next_url=nxt)

            student = Student.query.filter_by(school_id=school_id, nis=nis).first()
            
            if student and not student.is_active:
                flash("Akun Anda sudah dinonaktifkan. Hubungi admin lab sekolah.", "error")
            elif student and student.check_password(password):
                session.permanent = True
                session["student_id"] = student.id
                session["school_id"] = student.school_id
                session["student_password_version"] = student.password_version or 1
                flash(f"Selamat datang, {student.name} ({student.school_name})!", "success")
                return redirect(request.form.get("next") or nxt)
            else:
                flash("Sekolah, NIS, atau password salah.", "error")
                
        return render_template("login.html", schools=schools, next_url=nxt)

    @app.route("/student/dashboard")
    @student_required
    def student_dashboard():
        student = current_student()
        
        # Only fetch active and past borrowings for this specific student
        active = Borrowing.query.filter_by(
            student_id=student.id, status="active"
        ).order_by(Borrowing.borrow_date.desc()).all()
        past = Borrowing.query.filter(
            Borrowing.student_id == student.id,
            Borrowing.status.in_(["returned", "overdue"])
        ).order_by(Borrowing.borrow_date.desc()).limit(50).all()
        active_borrowing = active[0] if active else None
        
        return render_template(
            "student_dashboard.html",
            student=student,
            active_borrowing=active_borrowing,
            active_list=active,
            past=past,
            base_url=get_config().base_url
        )

    @app.route("/katalog")
    @student_required
    def student_catalog():
        student = current_student()
        school_id = student.school_id or 1
        
        tools = Tool.query.filter_by(school_id=school_id, is_active=True).order_by(Tool.code).all()
        total_alat = len(tools)
        tersedia = sum(1 for t in tools if t.is_available())
        dipinjam = sum(1 for t in tools if not t.is_available())
        categories = sorted(list(set(t.category for t in tools if t.category)))
        
        return render_template(
            "student_catalog.html",
            student=student,
            tools=tools,
            total_alat=total_alat,
            tersedia=tersedia,
            dipinjam=dipinjam,
            categories=categories,
            base_url=get_config().base_url
        )

    @app.route("/student/profile", methods=["GET", "POST"])
    @student_required
    def student_profile():
        student = current_student()
        if request.method == "POST":
            password_changed = False

            # 1. Password Update
            old_password = request.form.get("old_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")

            if old_password or new_password or confirm_password:
                if not student.check_password(old_password):
                    flash("Password lama yang Anda masukkan salah.", "error")
                    return redirect(url_for("student_profile"))
                if len(new_password) < 4:
                    flash("Password baru minimal 4 karakter.", "warning")
                    return redirect(url_for("student_profile"))
                if new_password != confirm_password:
                    flash("Konfirmasi password baru tidak cocok.", "error")
                    return redirect(url_for("student_profile"))
                
                student.set_password(new_password)
                student.password_version = (student.password_version or 1) + 1
                password_changed = True

            # 2. Profile Photo Avatar Upload
            if "avatar" in request.files:
                file = request.files["avatar"]
                if file and file.filename != "":
                    ext = os.path.splitext(file.filename)[1].lower()
                    if ext in [".jpg", ".jpeg", ".png", ".webp", ".gif"]:
                        os.makedirs("static/uploads/avatars", exist_ok=True)
                        filename = f"avatar_std_{student.id}_{int(utcnow().timestamp())}{ext}"
                        filepath = os.path.join("static/uploads/avatars", filename)
                        file.save(filepath)
                        student.avatar_path = filepath
                        if not password_changed:
                            flash("Foto profil berhasil diperbarui!", "success")
                    else:
                        flash("Format foto harus JPG, PNG, WEBP, atau GIF.", "error")

            db.session.commit()

            if password_changed:
                session.clear()
                flash("Password berhasil diperbarui! Demi keamanan, semua sesi login di seluruh perangkat telah di-logout. Silakan login kembali dengan password baru Anda.", "info")
                return redirect(url_for("student_login"))

            return redirect(url_for("student_profile"))

        return render_template("student_profile.html", student=student)

    @app.route("/logout")
    def logout():
        session.clear()
        flash("Anda sudah logout.", "info")
        return redirect(url_for("index"))

    @app.route("/clear-flash")
    def clear_flash():
        from flask import session as fl_session
        fl_session.pop("_flashes", None)
        return redirect(request.referrer or url_for("index"))

    @app.route("/pinjam/<code>", methods=["GET", "POST"])
    @student_required
    def pinjam(code):
        tool = Tool.query.filter_by(code=code, is_active=True).first_or_404()
        student = current_student()
        
        # 1. Enforce student ban check
        if student.is_banned:
            remaining = student.banned_remaining_str or "beberapa waktu"
            flash(f"Akun Anda sedang dibekukan ({remaining}) karena terdeteksi {student.spam_count}x percobaan peminjaman palsu/di luar area lab. Hubungi Admin Laboran untuk membuka blokir.", "error")
            return redirect(url_for("tool_detail", code=code))

        # 2. Enforce same school check
        if tool.school_id and student.school_id and tool.school_id != student.school_id:
            flash(f"Alat ini milik laboratorium {tool.school_name}. Anda terdaftar di {student.school_name}.", "error")
            return redirect(url_for("tool_detail", code=code))

        if not tool.is_available():
            flash(f"Maaf, {tool.name} sedang dipinjam.", "warning")
            return redirect(url_for("tool_detail", code=code))

        if request.method == "POST":
            notes = request.form.get("notes", "").strip()
            
            # Extract Location & Metadata from POST
            lat_str = request.form.get("lat")
            lng_str = request.form.get("lng")
            device_info = request.form.get("device_info", "").strip() or request.user_agent.string
            
            # Real Client IP extraction (handling Nginx X-Forwarded-For)
            ip_address = request.headers.get("X-Forwarded-For", request.remote_addr)
            if ip_address and "," in ip_address:
                ip_address = ip_address.split(",")[0].strip()

            borrow_lat = float(lat_str) if lat_str else None
            borrow_lng = float(lng_str) if lng_str else None

            # 3. Geofencing GPS Verification against School Location
            school = student.school or tool.school
            distance_m = None

            if school and school.require_geofence and school.latitude is not None and school.longitude is not None:
                if borrow_lat is not None and borrow_lng is not None:
                    from geofence import haversine_distance
                    distance_m = haversine_distance(borrow_lat, borrow_lng, school.latitude, school.longitude)
                    max_radius = school.max_geofence_radius_meters or 200

                    if distance_m > max_radius:
                        # Increment spam count for fake/out-of-range attempt
                        student.spam_count = (student.spam_count or 0) + 1
                        if student.spam_count >= 3:
                            student.banned_until = utcnow() + timedelta(days=7)
                            student.is_active = False

                        db.session.commit()

                        if student.spam_count >= 3:
                            flash(f"Peminjaman Ditolak! Posisi Anda terdeteksi {distance_m:.0f}m di luar area laboratorium (maksimal {max_radius}m). Akun Anda kini dibekukan 7 hari (3x percobaan palsu).", "error")
                        else:
                            flash(f"Peminjaman Ditolak! Posisi Anda terdeteksi {distance_m:.0f}m di luar area laboratorium (maksimal {max_radius}m). Teguran ke-{student.spam_count}/3.", "warning")

                        return redirect(url_for("tool_detail", code=code))
                else:
                    flash("Lokasi GPS (akses lokasi HP) wajib diizinkan untuk memvalidasi peminjaman di laboratorium sekolah.", "error")
                    return redirect(url_for("tool_detail", code=code))

            hours = Borrowing.default_deadline_hours(student.school_id if student else tool.school_id)
            new = Borrowing(
                school_id=student.school_id or tool.school_id,
                tool_id=tool.id,
                student_id=student.id,
                borrow_date=utcnow(),
                deadline=utcnow() + timedelta(hours=hours),
                notes=notes,
                borrow_lat=borrow_lat,
                borrow_lng=borrow_lng,
                borrow_distance_meters=distance_m,
                device_info=device_info[:255] if device_info else None,
                ip_address=ip_address[:50] if ip_address else None,
            )
            db.session.add(new)
            db.session.commit()
            flash(f"Berhasil meminjam {tool.name}. Kembalikan sebelum {new.deadline.strftime('%H:%M')}.", "success")
            return redirect(url_for("tool_detail", code=code))

        hours = Borrowing.default_deadline_hours(student.school_id if student else tool.school_id)
        school = student.school or tool.school
        return render_template("pinjam.html", tool=tool, duration_hours=hours, school=school)

    @app.route("/kembalikan/<code>", methods=["POST"])
    @student_required
    def kembalikan(code):
        tool = Tool.query.filter_by(code=code, is_active=True).first_or_404()
        student = current_student()
        current = tool.current_borrowing()
        if current is None:
            flash("Alat ini tidak sedang dipinjam.", "warning")
            return redirect(url_for("tool_detail", code=code))
        if current.student_id != student.id:
            flash("Alat ini dipinjam oleh siswa lain. Hubungi laboran jika perlu.", "error")
            return redirect(url_for("tool_detail", code=code))
        current.return_date = utcnow()
        current.status = "returned"
        current.condition_after = request.form.get("condition_after", "Baik")
        db.session.commit()
        flash(f"Terima kasih! {tool.name} sudah dikembalikan.", "success")
        return redirect(request.referrer or url_for("student_dashboard"))

    @app.route("/perpanjang/<int:borrowing_id>", methods=["GET", "POST"])
    @student_required
    def perpanjang(borrowing_id):
        borrowing = Borrowing.query.get_or_404(borrowing_id)
        student = current_student()
        if borrowing.student_id != student.id:
            flash("Anda tidak memiliki akses ke peminjaman ini.", "error")
            return redirect(url_for("student_dashboard"))
        if borrowing.status == "returned":
            flash("Peminjaman ini sudah selesai.", "warning")
            return redirect(url_for("student_dashboard"))
        
        max_extends = 2
        extend_count = getattr(borrowing, 'extend_count', 0)
        if extend_count >= max_extends:
            flash("Batas perpanjangan sudah tercapai (maksimal 2 kali).", "warning")
            return redirect(url_for("student_dashboard"))
        
        if request.method == "POST":
            choice = request.form.get("choice")
            if choice == "return":
                borrowing.return_date = utcnow()
                borrowing.status = "returned"
                db.session.commit()
                flash(f"{borrowing.tool.name} berhasil dikembalikan.", "success")
                return redirect(url_for("student_dashboard"))
            elif choice == "extend":
                hours = get_config().loan_duration_hours
                borrowing.deadline = borrowing.deadline + timedelta(hours=hours)
                borrowing.extend_count = extend_count + 1
                db.session.commit()
                new_deadline = borrowing.deadline.strftime('%d/%m/%Y %H:%M')
                flash(f"Peminjaman {borrowing.tool.name} diperpanjang hingga {new_deadline}.", "success")
                return redirect(url_for("student_dashboard"))
        
        can_extend = extend_count < max_extends
        return render_template(
            "extend_confirm.html",
            borrowing=borrowing,
            tool=borrowing.tool,
            loan_duration_hours=get_config().loan_duration_hours,
            can_extend=can_extend,
            extend_count=extend_count,
            max_extends=max_extends
        )

    @app.route("/riwayat")
    @student_required
    def my_history():
        student = current_student()
        borrowings = (Borrowing.query.filter_by(student_id=student.id)
                      .order_by(Borrowing.borrow_date.desc()).limit(50).all())
        return render_template("history.html", borrowings=borrowings)

    @app.route("/admin/login", methods=["GET", "POST"])
    def admin_login():
        if current_admin():
            return redirect(url_for("admin_dashboard"))
        nxt = request.args.get("next") or url_for("admin_dashboard")
        schools = School.query.filter_by(is_active=True).order_by(School.name).all()
        if request.method == "POST":
            school_id = request.form.get("school_id")
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            
            if not school_id or not username or not password:
                flash("Pilih asal sekolah, username, dan password wajib diisi.", "warning")
                return render_template("admin/admin_login.html", schools=schools, next_url=nxt)

            admin = Admin.query.filter_by(school_id=school_id, username=username).first()
            if not admin:
                admin = Admin.query.filter_by(username=username).first()

            if admin and admin.check_password(password):
                session.permanent = True
                session["admin_id"] = admin.id
                session["school_id"] = admin.school_id
                flash(f"Login berhasil. Halo, {admin.full_name or admin.username} ({admin.school_name}).", "success")
                return redirect(request.form.get("next") or nxt)
            flash("Sekolah, username, atau password salah.", "error")
        return render_template("admin/admin_login.html", schools=schools, next_url=nxt)

    @app.route("/admin/logout")
    def admin_logout():
        session.pop("admin_id", None)
        flash("Anda sudah logout.", "info")
        return redirect(url_for("index"))

    @app.route("/admin/dashboard")
    @admin_required
    def admin_dashboard():
        admin = current_admin()
        school_id = admin.school_id or 1
        
        active_query = Borrowing.query.filter_by(status="active")
        if school_id:
            active_query = active_query.filter_by(school_id=school_id)
        active = active_query.order_by(Borrowing.deadline.asc()).all()

        for b in active:
            b.seconds_left = b.seconds_remaining()
            b.elapsed_str = _humanize_seconds(b.elapsed_seconds())
            b.remaining_str = _humanize_seconds(b.seconds_left)

        tools_query = Tool.query.filter_by(is_active=True)
        if school_id:
            tools_query = tools_query.filter_by(school_id=school_id)
        all_tools = tools_query.all()

        total_alat = len(all_tools)
        tersedia = sum(1 for t in all_tools if t.is_available())
        dipinjam = sum(1 for t in all_tools if not t.is_available())
        telat = sum(1 for b in active if b.is_overdue)

        recent_query = Borrowing.query
        if school_id:
            recent_query = recent_query.filter_by(school_id=school_id)
        recent_borrowings = recent_query.order_by(Borrowing.borrow_date.desc()).limit(20).all()

        return render_template(
            "admin/dashboard.html", 
            total_alat=total_alat, 
            tersedia=tersedia, 
            dipinjam=dipinjam, 
            telat=telat,
            recent_borrowings=recent_borrowings,
            admin=admin
        )

    @app.route("/admin/tools")
    @admin_required
    def admin_tools():
        admin = current_admin()
        school_id = admin.school_id or 1
        search = request.args.get("search", "").strip()
        action = request.args.get("action", "")
        
        if action == "add":
            return render_template("admin/tool_form.html", tool=None, admin=admin)
        
        if action == "edit":
            code = request.args.get("code", "")
            tool = Tool.query.filter_by(code=code, school_id=school_id).first_or_404()
            return render_template("admin/tool_form.html", tool=tool, admin=admin)
        
        page = int(request.args.get("page", 1))
        per_page = 20
        query = Tool.query.filter_by(is_active=True, school_id=school_id)
        if search:
            query = query.filter(Tool.name.ilike(f"%{search}%") | Tool.code.ilike(f"%{search}%"))
        total = query.count()
        tools = query.order_by(Tool.code).offset((page - 1) * per_page).limit(per_page).all()
        total_pages = max(1, (total + per_page - 1) // per_page)
        return render_template("admin/tools.html", tools=tools, search=search, current_page=page, total_pages=total_pages, admin=admin)
    
    @app.route("/admin/tools/action", methods=["POST"])
    @admin_required
    def admin_tools_action():
        admin = current_admin()
        school_id = admin.school_id or 1
        action = request.form.get("_action", "add")
        
        if action == "add":
            code = request.form.get("code", "").strip().upper()
            name = request.form.get("name", "").strip()
            category = request.form.get("category", "").strip()
            lab_location = (request.form.get("lab_location") or request.form.get("location") or "").strip()
            condition = request.form.get("condition", "Baik")
            description = request.form.get("description", "").strip()
            icon = request.form.get("icon", "◇").strip()
            
            if not code or not name:
                flash("Kode dan nama alat wajib diisi.", "error")
                return redirect(url_for("admin_tools"))
            if Tool.query.filter_by(code=code, school_id=school_id).first():
                flash(f"Kode {code} sudah dipakai di sekolah ini.", "error")
                return redirect(url_for("admin_tools"))
            
            try:
                tool = Tool(
                    school_id=school_id, 
                    code=code, 
                    name=name, 
                    category=category, 
                    lab_location=lab_location, 
                    condition=condition, 
                    description=description, 
                    icon=icon
                )
                db.session.add(tool)
                db.session.commit()
                try:
                    qr_path = generate_qr_for_tool(tool)
                    tool.qr_path = qr_path
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    flash(f"Alat ditambahkan tapi QR gagal di-generate: {e}", "warning")
                flash(f"Alat {tool.name} ditambahkan.", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Gagal menambahkan alat: {e}", "error")
            return redirect(url_for("admin_tools"))
        
        elif action == "update":
            old_code = request.form.get("old_code", "").strip().upper()
            tool = Tool.query.filter_by(code=old_code, school_id=school_id).first_or_404()
            new_code = request.form.get("code", "").strip().upper()
            tool.code = new_code
            tool.name = request.form.get("name", "").strip()
            tool.category = request.form.get("category", "").strip()
            tool.lab_location = (request.form.get("lab_location") or request.form.get("location") or "").strip()
            tool.condition = request.form.get("condition", "Baik")
            tool.description = request.form.get("description", "").strip()
            tool.icon = request.form.get("icon", "◇").strip()
            db.session.commit()
            flash(f"Data {tool.name} diperbarui.", "success")
            return redirect(url_for("admin_tools"))
        
        elif action == "delete":
            code = request.form.get("code", "").strip().upper()
            if not code:
                code = request.form.get("id", "").strip().upper()
            tool = Tool.query.filter_by(code=code, school_id=school_id).first_or_404()
            if tool.current_borrowing():
                flash("Tidak bisa hapus alat yang sedang dipinjam.", "error")
            else:
                tool.is_active = False
                db.session.commit()
                flash(f"Alat {tool.name} dihapus.", "info")
            return redirect(url_for("admin_tools"))
        
        return redirect(url_for("admin_tools"))

    @app.route("/admin/generate-qr/<code>", methods=["POST"])
    @admin_required
    def admin_generate_qr(code):
        admin = current_admin()
        school_id = admin.school_id or 1
        tool = Tool.query.filter_by(code=code, school_id=school_id, is_active=True).first_or_404()
        path = generate_qr_for_tool(tool)
        tool.qr_path = path
        db.session.commit()
        return jsonify({"success": True, "qr_code": tool.qr_url})

    @app.route("/admin/generate-all-qr")
    @admin_required
    def admin_generate_all_qr():
        admin = current_admin()
        school_id = admin.school_id or 1
        tools = Tool.query.filter_by(school_id=school_id, is_active=True).all()
        for tool in tools:
            generate_qr_for_tool(tool)
        db.session.commit()
        flash(f"QR code berhasil di-generate untuk {len(tools)} alat.", "success")
        return redirect(url_for("admin_qr_labels"))

    @app.route("/admin/borrowings")
    @admin_required
    def admin_borrowings():
        admin = current_admin()
        school_id = admin.school_id or 1
        search = request.args.get("search", "").strip()
        status_filter = request.args.get("status_filter", "").strip()
        
        q = Borrowing.query.filter_by(school_id=school_id).order_by(Borrowing.borrow_date.desc())
        if search:
            q = q.filter(
                db.or_(
                    Borrowing.student_name.ilike(f"%{search}%"),
                    Borrowing.tool_name.ilike(f"%{search}%"),
                    Borrowing.tool_code.ilike(f"%{search}%"),
                )
            )
        if status_filter == "active":
            q = q.filter_by(status="active")
        elif status_filter == "returned":
            q = q.filter_by(status="returned")
        elif status_filter == "overdue":
            q = q.filter_by(status="active").filter(Borrowing.deadline < utcnow())
        borrowings = q.limit(100).all()
        for b in borrowings:
            b.seconds_left = b.seconds_remaining()
        return render_template("admin/borrowings.html", borrowings=borrowings, 
                               search=search, status_filter=status_filter, admin=admin)

    @app.route("/admin/borrowings/export-csv")
    @admin_required
    def admin_export_borrowings_csv():
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        from flask import send_file

        admin = current_admin()
        school_id = admin.school_id or 1
        search = request.args.get("search", "").strip()
        status_filter = request.args.get("status_filter", "").strip()

        q = Borrowing.query.filter_by(school_id=school_id).order_by(Borrowing.borrow_date.desc())
        if search:
            q = q.filter(
                db.or_(
                    Borrowing.student_name.ilike(f"%{search}%"),
                    Borrowing.tool_name.ilike(f"%{search}%"),
                    Borrowing.tool_code.ilike(f"%{search}%"),
                )
            )
        if status_filter == "active":
            q = q.filter_by(status="active")
        elif status_filter == "returned":
            q = q.filter_by(status="returned")
        elif status_filter == "overdue":
            q = q.filter_by(status="active").filter(Borrowing.deadline < utcnow())

        borrowings = q.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Laporan Peminjaman"
        ws.views.sheetView[0].showGridLines = True

        title_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        subtitle_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        even_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        status_returned_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        status_returned_font = Font(name="Segoe UI", size=10, bold=True, color="15803D")

        status_overdue_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        status_overdue_font = Font(name="Segoe UI", size=10, bold=True, color="B91C1C")

        status_active_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        status_active_font = Font(name="Segoe UI", size=10, bold=True, color="1D4ED8")

        title_font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
        subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="E2E8F0")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        body_font = Font(name="Segoe UI", size=10, color="0F172A")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        ws.merge_cells("A1:K1")
        ws["A1"] = f"LAPORAN REKAPITULASI PEMINJAMAN ALAT — {admin.school_name.upper()}"
        ws["A1"].font = title_font
        ws["A1"].fill = title_fill
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 32

        ws.merge_cells("A2:K2")
        now_str = utcnow().strftime("%d/%m/%Y %H:%M WIB")
        ws["A2"] = f"Sistem LabKeeper — {admin.school_name}  |  Tanggal Cetak: {now_str}  |  Total Data: {len(borrowings)}"
        ws["A2"].font = subtitle_font
        ws["A2"].fill = subtitle_fill
        ws["A2"].alignment = align_center
        ws.row_dimensions[2].height = 22

        ws.row_dimensions[3].height = 10

        headers = ["No", "ID", "Nama Siswa", "NIS", "Kode Alat", "Nama Alat", "Tgl Pinjam", "Batas Kembali", "Tgl Kembali", "Status", "Kondisi Akhir"]
        ws.row_dimensions[4].height = 26

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

        start_row = 5
        for idx, b in enumerate(borrowings, 1):
            row_num = start_row + idx - 1
            ws.row_dimensions[row_num].height = 22
            row_fill = even_row_fill if idx % 2 == 0 else odd_row_fill

            is_returned = bool(b.return_time)
            is_overdue = b.is_overdue() if hasattr(b, 'is_overdue') else False
            status_str = "Dikembalikan" if is_returned else ("Telat" if is_overdue else "Aktif")

            row_data = [
                (idx, align_center),
                (f"#{b.id}", align_center),
                (b.student_name or "-", align_left),
                (b.student_nis or "-", align_center),
                (b.tool_code or "-", align_center),
                (b.tool_name or "-", align_left),
                (b.start_time.strftime("%d/%m/%Y %H:%M") if b.start_time else "-", align_center),
                (b.due_time.strftime("%d/%m/%Y %H:%M") if b.due_time else "-", align_center),
                (b.return_time.strftime("%d/%m/%Y %H:%M") if b.return_time else "-", align_center),
                (status_str, align_center),
                (b.condition_after or "Baik", align_center),
            ]

            for col_idx, (val, alignment) in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = body_font
                cell.fill = row_fill
                cell.alignment = alignment
                cell.border = thin_border

                if col_idx == 10:
                    if status_str == "Dikembalikan":
                        cell.fill = status_returned_fill
                        cell.font = status_returned_font
                    elif status_str == "Telat":
                        cell.fill = status_overdue_fill
                        cell.font = status_overdue_font
                    else:
                        cell.fill = status_active_fill
                        cell.font = status_active_font

        padding = {1: 6, 2: 8, 3: 24, 4: 14, 5: 14, 6: 24, 7: 20, 8: 20, 9: 20, 10: 16, 11: 16}
        for col_idx, width in padding.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"laporan_peminjaman_{utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    @app.route("/admin/return/<int:bid>", methods=["GET", "POST"])
    @admin_required
    def admin_return(bid):
        admin = current_admin()
        school_id = admin.school_id or 1
        b = Borrowing.query.filter_by(id=bid, school_id=school_id).first_or_404()
        if request.method == "POST":
            b.return_date = utcnow()
            b.status = "returned"
            b.condition_after = request.form.get("condition_after", "Baik")
            db.session.commit()
            flash(f"{b.tool_name} dikembalikan oleh {b.student_name}.", "success")
            return redirect(request.referrer or url_for("admin_borrowings"))
        return redirect(request.referrer or url_for("admin_borrowings"))

    @app.route("/admin/extend/<int:bid>", methods=["GET", "POST"])
    @admin_required
    def admin_extend(bid):
        admin = current_admin()
        school_id = admin.school_id or 1
        b = Borrowing.query.filter_by(id=bid, school_id=school_id).first_or_404()
        if b.status == "returned":
            flash("Peminjaman sudah dikembalikan.", "warning")
            return redirect(url_for("admin_borrowings"))
        try:
            hours = int(request.form.get("hours", 2)) if request.method == "POST" else int(request.args.get("hours", 2))
        except (TypeError, ValueError):
            hours = 2
        hours = max(1, min(hours, 168))

        base = b.deadline if b.deadline > utcnow() else utcnow()
        b.deadline = base + timedelta(hours=hours)
        b.extend_count = (b.extend_count or 0) + 1
        if b.status == "overdue":
            b.status = "active"
        db.session.commit()
        flash(f"Deadline {b.tool_name} diperpanjang admin +{hours} jam → {b.deadline.strftime('%d/%m/%Y %H:%M')}.", "success")
        return redirect(url_for("admin_borrowings"))

    @app.route("/admin/borrowings/bulk-action", methods=["POST"])
    @admin_required
    def admin_borrowings_bulk_action():
        admin = current_admin()
        school_id = admin.school_id or 1
        action = request.form.get("_action", "")
        bid_list = request.form.getlist("bid")
        try:
            hours = max(1, min(int(request.form.get("hours", 2)), 168))
        except (TypeError, ValueError):
            hours = 2

        if not bid_list:
            flash("Pilih minimal satu peminjaman terlebih dahulu.", "warning")
            return redirect(url_for("admin_borrowings"))

        bid_ints = [int(x) for x in bid_list if x.isdigit()]
        borrowings = Borrowing.query.filter(Borrowing.id.in_(bid_ints), Borrowing.school_id == school_id).all()
        if not borrowings:
            flash("Tidak ada peminjaman yang cocok.", "error")
            return redirect(url_for("admin_borrowings"))

        if action == "return":
            count = 0
            for b in borrowings:
                if b.status != "returned":
                    b.return_date = utcnow()
                    b.status = "returned"
                    b.condition_after = request.form.get("condition_after", "Baik")
                    count += 1
            db.session.commit()
            flash(f"{count} peminjaman ditandai dikembalikan.", "success")

        elif action == "extend":
            count = 0
            for b in borrowings:
                if b.status == "returned":
                    continue
                base = b.deadline if b.deadline > utcnow() else utcnow()
                b.deadline = base + timedelta(hours=hours)
                b.extend_count = (b.extend_count or 0) + 1
                if b.status == "overdue":
                    b.status = "active"
                count += 1
            db.session.commit()
            flash(f"{count} peminjaman diperpanjang +{hours} jam.", "success")
        else:
            flash(f"Aksi tidak dikenal: {action}", "error")

        return redirect(url_for("admin_borrowings"))

    @app.route("/admin/students")
    @admin_required
    def admin_students():
        admin = current_admin()
        school_id = admin.school_id or 1
        search = request.args.get("search", "").strip()
        status_filter = request.args.get("status_filter", "").strip()
        action = request.args.get("action", "")
        
        if action == "add":
            return render_template("admin/student_form.html", student=None, admin=admin)
        
        if action == "edit":
            nis = request.args.get("nis", "")
            student = Student.query.filter_by(nis=nis, school_id=school_id).first_or_404()
            return render_template("admin/student_form.html", student=student, admin=admin)
        
        query = Student.query.filter_by(school_id=school_id)
        if status_filter == "active":
            query = query.filter_by(is_active=True)
        elif status_filter == "inactive":
            query = query.filter_by(is_active=False)

        if search:
            query = query.filter(
                db.or_(
                    Student.name.ilike(f"%{search}%"),
                    Student.nis.ilike(f"%{search}%"),
                    Student.class_name.ilike(f"%{search}%"),
                )
            )
        students = query.order_by(Student.class_name, Student.name).all()
        return render_template("admin/students.html", students=students, search=search, status_filter=status_filter, admin=admin)

    @app.route("/admin/students/action", methods=["POST"])
    @admin_required
    def admin_students_action():
        admin = current_admin()
        school_id = admin.school_id or 1
        action = request.form.get("_action", "add")
        
        if action == "add":
            nis = request.form.get("nis", "").strip()
            name = request.form.get("name", "").strip()
            class_name = request.form.get("class", "").strip()
            password = request.form.get("password", "").strip()
            
            if not nis or not name or not class_name or not password:
                flash("Semua field wajib diisi.", "error")
                return redirect(url_for("admin_students"))
            if Student.query.filter_by(school_id=school_id, nis=nis).first():
                flash(f"NIS {nis} sudah terdaftar di sekolah ini.", "error")
                return redirect(url_for("admin_students"))
            try:
                s = Student(school_id=school_id, nis=nis, name=name, class_name=class_name)
                s.set_password(password)
                db.session.add(s)
                db.session.commit()
                flash(f"Siswa {s.name} ditambahkan.", "success")
            except Exception as e:
                db.session.rollback()
                flash(f"Gagal menambahkan siswa: {e}", "error")
            return redirect(url_for("admin_students"))
        
        elif action == "update":
            old_nis = request.form.get("old_nis", "").strip()
            s = Student.query.filter_by(school_id=school_id, nis=old_nis).first_or_404()
            new_nis = request.form.get("nis", "").strip()
            s.nis = new_nis
            s.name = request.form.get("name", "").strip()
            s.class_name = request.form.get("class", "").strip()
            pw = request.form.get("password", "").strip()
            if pw:
                s.set_password(pw)
            db.session.commit()
            flash(f"Siswa {s.name} diperbarui.", "success")
            return redirect(url_for("admin_students"))

        elif action == "delete":
            nis = request.form.get("nis", "").strip()
            s = Student.query.filter_by(school_id=school_id, nis=nis).first_or_404()

            Borrowing.query.filter_by(student_id=s.id).update({
                Borrowing.student_id: None,
                Borrowing.archived_student_name: s.name,
                Borrowing.archived_student_nis: s.nis,
            })

            db.session.delete(s)
            db.session.commit()
            flash(f"Siswa {s.name} dihapus permanen. Histori peminjaman tetap tersimpan.", "info")
            return redirect(url_for("admin_students"))

        elif action == "unblock":
            nis = request.form.get("nis", "").strip()
            s = Student.query.filter_by(school_id=school_id, nis=nis).first_or_404()
            s.is_active = True
            s.banned_until = None
            s.spam_count = 0
            db.session.commit()
            flash(f"Akun siswa {s.name} (NIS: {s.nis}) telah dibuka blokirnya dan jumlah teguran telah di-reset ke 0.", "success")
            return redirect(url_for("admin_students"))
        
        return redirect(url_for("admin_students"))

    @app.route("/admin/students/bulk-action", methods=["POST"])
    @admin_required
    def admin_students_bulk_action():
        admin = current_admin()
        school_id = admin.school_id or 1
        action = request.form.get("_action", "")
        nis_list = request.form.getlist("nis")

        if not nis_list:
            flash("Pilih minimal satu siswa terlebih dahulu.", "warning")
            return redirect(url_for("admin_students"))

        students = Student.query.filter(Student.nis.in_(nis_list), Student.school_id == school_id).all()
        if not students:
            flash("Tidak ada siswa yang cocok dengan pilihan.", "error")
            return redirect(url_for("admin_students"))

        if action == "activate":
            for s in students:
                s.is_active = True
            db.session.commit()
            flash(f"{len(students)} siswa diaktifkan.", "success")

        elif action == "deactivate":
            for s in students:
                s.is_active = False
            db.session.commit()
            flash(f"{len(students)} siswa dinonaktifkan.", "info")

        elif action == "delete":
            for s in students:
                Borrowing.query.filter_by(student_id=s.id).update({
                    Borrowing.student_id: None,
                    Borrowing.archived_student_name: s.name,
                    Borrowing.archived_student_nis: s.nis,
                })
                db.session.delete(s)
            db.session.commit()
            flash(f"{len(students)} siswa dihapus permanen. Histori peminjaman tetap tersimpan.", "info")
        return redirect(url_for("admin_students"))

    @app.route("/admin/students/export-excel")
    @admin_required
    def admin_students_export_excel():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        admin = current_admin()
        school_id = admin.school_id or 1
        search = request.args.get("search", "").strip()
        status_filter = request.args.get("status_filter", "").strip()

        query = Student.query.filter_by(school_id=school_id)
        if status_filter == "active":
            query = query.filter_by(is_active=True)
        elif status_filter == "inactive":
            query = query.filter_by(is_active=False)
        
        if search:
            query = query.filter(
                Student.name.ilike(f"%{search}%") | 
                Student.nis.ilike(f"%{search}%") | 
                Student.class_name.ilike(f"%{search}%")
            )
        
        students = query.order_by(Student.class_name, Student.nis).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Daftar Siswa"

        ws.merge_cells("A1:E1")
        title_cell = ws["A1"]
        title_cell.value = f"DAFTAR REKAPITULASI SISWA TERDAFTAR — {admin.school_name.upper()}"
        title_cell.font = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        now_str = utcnow().strftime("%d/%m/%Y %H:%M WIB")
        ws.merge_cells("A2:E2")
        sub_cell = ws["A2"]
        sub_cell.value = f"Dicetak pada: {now_str}  |  Total Siswa: {len(students)}"
        sub_cell.font = Font(name="Segoe UI", size=9, italic=True, color="475569")
        sub_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        ws.row_dimensions[3].height = 6

        headers = ["No", "NIS", "Nama Lengkap Siswa", "Kelas", "Status Akun"]
        ws.row_dimensions[4].height = 26
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="F8FAFC")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        data_font = Font(name="Segoe UI", size=10, color="0F172A")

        for idx, s in enumerate(students, 1):
            row_num = 4 + idx
            ws.row_dimensions[row_num].height = 22
            row_vals = [idx, s.nis, s.name, s.class_name, "Aktif" if s.is_active else "Nonaktif"]
            for col_idx, val in enumerate(row_vals, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.fill = row_fill
                if col_idx in [1, 2, 4, 5]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 16

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"Daftar_Siswa_{admin.school_name.replace(' ', '_')}.xlsx")

    @app.route("/admin/students/download-template")
    @admin_required
    def admin_students_download_template():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Data Siswa"
        
        # Title Banner
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "TEMPLATE IMPORT DATA SISWA — LABKEEPER"
        title_cell.font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # Subtitle Instruction Banner
        ws.merge_cells("A2:D2")
        sub_cell = ws["A2"]
        sub_cell.value = "Petunjuk: Isi data siswa baru mulai dari baris ke-4. Tanda (*) menunjukkan kolom wajib."
        sub_cell.font = Font(name="Segoe UI", size=9, italic=True, color="475569")
        sub_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        ws.row_dimensions[3].height = 6

        # Table Header
        headers = ["NIS*", "Nama Lengkap Siswa*", "Kelas*", "Password (Opsional)"]
        ws.row_dimensions[4].height = 26
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="F8FAFC")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # CLEAN TEMPLATE - Clean empty formatted rows ready to type!
        row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(name="Segoe UI", size=10, color="0F172A")
        
        for r_idx in range(5, 25):
            ws.row_dimensions[r_idx].height = 22
            for c_idx in range(1, 5):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.fill = row_fill

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 24

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="Template_Import_Siswa_LabKeeper.xlsx")

    @app.route("/admin/students/preview-excel", methods=["POST"])
    @admin_required
    def admin_students_preview_excel():
        admin = current_admin()
        school_id = admin.school_id or 1

        if "file" not in request.files:
            return jsonify({"success": False, "message": "Pilih file Excel (.xlsx) terlebih dahulu."}), 400
        
        file = request.files["file"]
        if not file or not file.filename.endswith(".xlsx"):
            return jsonify({"success": False, "message": "Format file harus .xlsx (Excel)."}), 400

        from openpyxl import load_workbook
        import uuid
        try:
            wb = load_workbook(filename=file, data_only=True)
            ws = wb.active

            first_rows = list(ws.iter_rows(values_only=True))[:6]
            first_rows_text = str([[cell for cell in row if cell is not None] for row in first_rows])
            
            if "Kode Alat" in first_rows_text or "Lokasi Rak" in first_rows_text:
                return jsonify({"success": False, "message": "Gagal Impor: File yang diunggah adalah Template Alat Lab, bukan Template Siswa! Harap gunakan file Template Import Siswa."}), 400

            has_nis = any("NIS" in str(cell).upper() for row in first_rows for cell in row if cell is not None)
            has_name = any("NAMA" in str(cell).upper() for row in first_rows for cell in row if cell is not None)

            if not (has_nis and has_name):
                return jsonify({"success": False, "message": "Gagal Impor: File Excel yang diunggah tidak dikenali sebagai Template Siswa LabKeeper."}), 400

            temp_dir = os.path.join(app.static_folder, "uploads", "temp")
            os.makedirs(temp_dir, exist_ok=True)
            temp_filename = f"temp_students_{uuid.uuid4().hex}.xlsx"
            temp_filepath = os.path.join(temp_dir, temp_filename)
            file.seek(0)
            file.save(temp_filepath)

            parsed_rows = []
            valid_count = 0
            duplicate_count = 0
            invalid_count = 0

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if not row or not any(row):
                    continue
                
                row_str = " ".join([str(c) for c in row if c is not None])
                if "TEMPLATE" in row_str or "Petunjuk" in row_str or "NIS*" in row_str or "Nama Lengkap" in row_str or str(row[0]).strip().upper() in ["NIS", "NIS*"]:
                    continue

                nis = str(row[0]).strip() if row[0] is not None else ""
                name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                class_name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                password = str(row[3]).strip() if len(row) > 3 and row[3] is not None else "123456"

                if not nis and not name and not class_name:
                    continue

                if not nis or not name or not class_name:
                    invalid_count += 1
                    parsed_rows.append({
                        "row_idx": i + 1, "nis": nis or "-", "name": name or "-", "class_name": class_name or "-",
                        "status": "invalid", "status_label": "Data Tidak Lengkap (Dilewati)"
                    })
                    continue

                if Student.query.filter_by(school_id=school_id, nis=nis).first():
                    duplicate_count += 1
                    parsed_rows.append({
                        "row_idx": i + 1, "nis": nis, "name": name, "class_name": class_name,
                        "status": "duplicate", "status_label": "NIS Sudah Ada (Dilewati)"
                    })
                    continue

                valid_count += 1
                parsed_rows.append({
                    "row_idx": i + 1, "nis": nis, "name": name, "class_name": class_name,
                    "status": "valid", "status_label": "Siap Di-import"
                })

            if not parsed_rows:
                if os.path.exists(temp_filepath):
                    os.remove(temp_filepath)
                return jsonify({"success": False, "message": "File Excel masih kosong atau tidak berisi data siswa baru."}), 400

            return jsonify({
                "success": True,
                "temp_file_id": temp_filename,
                "total_rows": len(parsed_rows),
                "valid_count": valid_count,
                "duplicate_count": duplicate_count,
                "invalid_count": invalid_count,
                "rows": parsed_rows
            })
        except Exception as e:
            return jsonify({"success": False, "message": f"Gagal membaca file Excel: {e}"}), 400

    @app.route("/admin/students/confirm-import", methods=["POST"])
    @admin_required
    def admin_students_confirm_import():
        admin = current_admin()
        school_id = admin.school_id or 1
        temp_file_id = request.form.get("temp_file_id", "").strip()

        if not temp_file_id:
            flash("ID file temporary tidak ditemukan.", "error")
            return redirect(url_for("admin_students"))

        temp_filepath = os.path.join(app.static_folder, "uploads", "temp", temp_file_id)
        if not os.path.exists(temp_filepath):
            flash("File preview telah kadaluarsa. Silakan upload ulang.", "error")
            return redirect(url_for("admin_students"))

        from openpyxl import load_workbook
        try:
            wb = load_workbook(filename=temp_filepath, data_only=True)
            ws = wb.active
            count_added = 0

            for row in ws.iter_rows(values_only=True):
                if not row or not any(row):
                    continue
                
                row_str = " ".join([str(c) for c in row if c is not None])
                if "TEMPLATE" in row_str or "Petunjuk" in row_str or "NIS*" in row_str or "Nama Lengkap" in row_str or str(row[0]).strip().upper() in ["NIS", "NIS*"]:
                    continue

                nis = str(row[0]).strip() if row[0] is not None else ""
                name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                class_name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
                password = str(row[3]).strip() if len(row) > 3 and row[3] is not None else "123456"

                if not nis or not name or not class_name:
                    continue

                if Student.query.filter_by(school_id=school_id, nis=nis).first():
                    continue

                s = Student(school_id=school_id, nis=nis, name=name, class_name=class_name)
                s.set_password(password if password else "123456")
                db.session.add(s)
                count_added += 1

            db.session.commit()
            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)
            flash(f"Berhasil mengimpor {count_added} siswa baru ke database!", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Gagal mengimpor data: {e}", "error")

        return redirect(url_for("admin_students"))

    @app.route("/admin/tools/download-template")
    @admin_required
    def admin_tools_download_template():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Data Alat Lab"
        
        # Title Banner
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "TEMPLATE IMPORT DATA ALAT LAB — LABKEEPER"
        title_cell.font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # Subtitle Instruction Banner
        ws.merge_cells("A2:F2")
        sub_cell = ws["A2"]
        sub_cell.value = "Petunjuk: Isi data alat lab baru mulai dari baris ke-4. Tanda (*) menunjukkan kolom wajib. QR Code akan dibuat otomatis."
        sub_cell.font = Font(name="Segoe UI", size=9, italic=True, color="475569")
        sub_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        sub_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        ws.row_dimensions[3].height = 6

        # Table Header
        headers = ["Kode Alat*", "Nama Alat*", "Kategori", "Lokasi Rak", "Kondisi", "Deskripsi"]
        ws.row_dimensions[4].height = 26
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="F8FAFC")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # CLEAN TEMPLATE - Clean empty formatted rows ready to type!
        row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        data_font = Font(name="Segoe UI", size=10, color="0F172A")
        
        for r_idx in range(5, 25):
            ws.row_dimensions[r_idx].height = 22
            for c_idx in range(1, 7):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = data_font
                cell.border = thin_border
                cell.fill = row_fill

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 32

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="Template_Import_Alat_LabKeeper.xlsx")

    @app.route("/admin/tools/preview-excel", methods=["POST"])
    @admin_required
    def admin_tools_preview_excel():
        admin = current_admin()
        school_id = admin.school_id or 1

        if "file" not in request.files:
            return jsonify({"success": False, "message": "Pilih file Excel (.xlsx) terlebih dahulu."}), 400
        
        file = request.files["file"]
        if not file or not file.filename.endswith(".xlsx"):
            return jsonify({"success": False, "message": "Format file harus .xlsx (Excel)."}), 400

        from openpyxl import load_workbook
        import uuid
        try:
            wb = load_workbook(filename=file, data_only=True)
            ws = wb.active

            first_rows = list(ws.iter_rows(values_only=True))[:6]
            first_rows_text = str([[cell for cell in row if cell is not None] for row in first_rows])
            
            if "NIS" in first_rows_text or "Nama Lengkap Siswa" in first_rows_text:
                return jsonify({"success": False, "message": "Gagal Impor: File yang diunggah adalah Template Siswa, bukan Template Alat Lab! Harap gunakan file Template Import Alat."}), 400

            has_code = any("KODE" in str(cell).upper() for row in first_rows for cell in row if cell is not None)
            has_name = any("NAMA" in str(cell).upper() for row in first_rows for cell in row if cell is not None)

            if not (has_code and has_name):
                return jsonify({"success": False, "message": "Gagal Impor: File Excel yang diunggah tidak dikenali sebagai Template Alat LabKeeper."}), 400

            temp_dir = os.path.join(app.static_folder, "uploads", "temp")
            os.makedirs(temp_dir, exist_ok=True)
            temp_filename = f"temp_tools_{uuid.uuid4().hex}.xlsx"
            temp_filepath = os.path.join(temp_dir, temp_filename)
            file.seek(0)
            file.save(temp_filepath)

            parsed_rows = []
            valid_count = 0
            duplicate_count = 0
            invalid_count = 0

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if not row or not any(row):
                    continue
                
                row_str = " ".join([str(c) for c in row if c is not None])
                if "TEMPLATE" in row_str or "Petunjuk" in row_str or "Kode Alat*" in row_str or "Nama Alat*" in row_str or str(row[0]).strip().upper() in ["KODE ALAT", "KODE ALAT*"]:
                    continue

                code = str(row[0]).strip().upper() if row[0] is not None else ""
                name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                category = str(row[2]).strip() if len(row) > 2 and row[2] is not None else "Lainnya"
                lab_location = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                condition = str(row[4]).strip() if len(row) > 4 and row[4] is not None else "Baik"

                if not code and not name:
                    continue

                if not code or not name:
                    invalid_count += 1
                    parsed_rows.append({
                        "row_idx": i + 1, "code": code or "-", "name": name or "-", "category": category, "lab_location": lab_location or "-",
                        "status": "invalid", "status_label": "Data Tidak Lengkap (Dilewati)"
                    })
                    continue

                if Tool.query.filter_by(school_id=school_id, code=code).first():
                    duplicate_count += 1
                    parsed_rows.append({
                        "row_idx": i + 1, "code": code, "name": name, "category": category, "lab_location": lab_location or "-",
                        "status": "duplicate", "status_label": "Kode Sudah Ada (Dilewati)"
                    })
                    continue

                valid_count += 1
                parsed_rows.append({
                    "row_idx": i + 1, "code": code, "name": name, "category": category, "lab_location": lab_location or "-",
                    "status": "valid", "status_label": "Siap Di-import & Generate QR"
                })

            if not parsed_rows:
                if os.path.exists(temp_filepath):
                    os.remove(temp_filepath)
                return jsonify({"success": False, "message": "File Excel masih kosong atau tidak berisi data alat baru."}), 400

            return jsonify({
                "success": True,
                "temp_file_id": temp_filename,
                "total_rows": len(parsed_rows),
                "valid_count": valid_count,
                "duplicate_count": duplicate_count,
                "invalid_count": invalid_count,
                "rows": parsed_rows
            })
        except Exception as e:
            return jsonify({"success": False, "message": f"Gagal membaca file Excel: {e}"}), 400

    @app.route("/admin/tools/confirm-import", methods=["POST"])
    @admin_required
    def admin_tools_confirm_import():
        admin = current_admin()
        school_id = admin.school_id or 1
        temp_file_id = request.form.get("temp_file_id", "").strip()

        if not temp_file_id:
            flash("ID file temporary tidak ditemukan.", "error")
            return redirect(url_for("admin_tools"))

        temp_filepath = os.path.join(app.static_folder, "uploads", "temp", temp_file_id)
        if not os.path.exists(temp_filepath):
            flash("File preview telah kadaluarsa. Silakan upload ulang.", "error")
            return redirect(url_for("admin_tools"))

        from openpyxl import load_workbook
        try:
            wb = load_workbook(filename=temp_filepath, data_only=True)
            ws = wb.active
            count_added = 0

            for row in ws.iter_rows(values_only=True):
                if not row or not any(row):
                    continue
                
                row_str = " ".join([str(c) for c in row if c is not None])
                if "TEMPLATE" in row_str or "Petunjuk" in row_str or "Kode Alat*" in row_str or "Nama Alat*" in row_str or str(row[0]).strip().upper() in ["KODE ALAT", "KODE ALAT*"]:
                    continue

                code = str(row[0]).strip().upper() if row[0] is not None else ""
                name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                category = str(row[2]).strip() if len(row) > 2 and row[2] is not None else "Lainnya"
                lab_location = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
                condition = str(row[4]).strip() if len(row) > 4 and row[4] is not None else "Baik"
                description = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ""

                if not code or not name:
                    continue

                if Tool.query.filter_by(school_id=school_id, code=code).first():
                    continue

                t = Tool(
                    school_id=school_id,
                    code=code,
                    name=name,
                    category=category,
                    lab_location=lab_location,
                    condition=condition if condition in ["Baik", "Rusak Ringan", "Rusak Berat", "Perlu Perbaikan"] else "Baik",
                    description=description,
                    icon="◇"
                )
                db.session.add(t)
                db.session.commit()

                try:
                    t.qr_path = generate_qr_for_tool(t)
                    db.session.commit()
                except Exception:
                    pass

                count_added += 1

            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)
            flash(f"Berhasil mengimpor {count_added} alat lab baru dan generate QR otomatis!", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Gagal mengimpor data: {e}", "error")

        return redirect(url_for("admin_tools"))

    @app.route("/admin/qr-labels")
    @admin_required
    def admin_qr_labels():
        admin = current_admin()
        school_id = admin.school_id or 1
        tools = Tool.query.filter_by(school_id=school_id, is_active=True).order_by(Tool.code).all()
        return render_template("admin/qr_labels.html", tools=tools, admin=admin)

    @app.route("/admin/settings", methods=["GET", "POST"])
    @admin_required
    def admin_settings():
        admin = current_admin()
        school_id = admin.school_id or 1
        school = School.query.get(school_id)
        cfg = Config.get_solo()
        if request.method == "POST":
            # 1. Update Durasi Peminjaman Khusus Sekolah Ini
            duration = request.form.get("duration_hours")
            if duration:
                try:
                    val = max(1, int(duration))
                    if school:
                        school.loan_duration_hours = val
                    if cfg and school_id == 1:
                        cfg.loan_duration_hours = val
                except ValueError:
                    pass

            # 2. Update Pengaturan GPS Geofencing Sekolah
            lat_val = request.form.get("latitude", "").strip()
            lng_val = request.form.get("longitude", "").strip()
            radius_val = request.form.get("max_geofence_radius_meters", "").strip()
            require_geofence = request.form.get("require_geofence") == "1"

            if school:
                try:
                    school.latitude = float(lat_val) if lat_val else None
                except ValueError:
                    pass
                try:
                    school.longitude = float(lng_val) if lng_val else None
                except ValueError:
                    pass
                try:
                    if radius_val:
                        school.max_geofence_radius_meters = max(10, int(radius_val))
                except ValueError:
                    pass
                school.require_geofence = require_geofence

            # 3. Update Password Admin (jika diisi)
            current_pw = request.form.get("current_password", "").strip()
            new_pw = request.form.get("new_password", "").strip()
            confirm_pw = request.form.get("confirm_password", "").strip()

            if new_pw or current_pw or confirm_pw:
                if not current_pw:
                    flash("Masukkan password lama Anda untuk mengubah password.", "error")
                    return redirect(url_for("admin_settings"))
                if not admin.check_password(current_pw):
                    flash("Password lama yang Anda masukkan salah.", "error")
                    return redirect(url_for("admin_settings"))
                if not new_pw:
                    flash("Password baru tidak boleh kosong.", "warning")
                    return redirect(url_for("admin_settings"))
                if len(new_pw) < 6:
                    flash("Password baru minimal 6 karakter.", "warning")
                    return redirect(url_for("admin_settings"))
                if new_pw != confirm_pw:
                    flash("Konfirmasi password baru tidak cocok.", "warning")
                    return redirect(url_for("admin_settings"))

                admin.set_password(new_pw)
                flash("Password Admin berhasil diperbarui!", "success")

            db.session.commit()
            flash("Pengaturan berhasil disimpan.", "success")
            return redirect(url_for("admin_settings"))

        return render_template("admin/settings.html", settings=cfg, admin=admin, school=school)

    @app.route("/admin/reset-borrowings")
    @admin_required
    def admin_reset_borrowings():
        admin = current_admin()
        school_id = admin.school_id or 1
        Borrowing.query.filter_by(school_id=school_id).delete()
        db.session.commit()
        flash(f"Semua riwayat peminjaman {admin.school_name} direset.", "info")
        return redirect(url_for("admin_settings"))

    # ============= JSON API (for live countdown) =============
    @app.route("/api/active-borrowings")
    @admin_required
    def api_active_borrowings():
        admin = current_admin()
        school_id = admin.school_id or 1
        active = Borrowing.query.filter_by(school_id=school_id, status="active").all()
        result = []
        for b in active:
            result.append({
                "id": b.id,
                "tool_code": b.tool.code,
                "tool_name": b.tool.name,
                "student_name": b.student.name if b.student else (b.archived_student_name or "-"),
                "student_nis": b.student.nis if b.student else (b.archived_student_nis or "-"),
                "student_class": b.student.class_name if b.student else "-",
                "elapsed_seconds": b.elapsed_seconds(),
                "remaining_seconds": b.seconds_remaining(),
                "deadline_iso": b.deadline.isoformat() + "Z",
            })
        return jsonify(result)

    return app


def _humanize_seconds(seconds: int) -> str:
    """Convert seconds to a short Indonesian-friendly string."""
    sign = ""
    if seconds < 0:
        sign = "-"
        seconds = abs(seconds)
    h, rem = divmod(int(seconds), 3600)
    m, s = divmod(rem, 60)
    if h > 0:
        return f"{sign}{h}j {m}m"
    if m > 0:
        return f"{sign}{m}m {s}d"
    return f"{sign}{s}d"


# Expose the app for `flask run`
app = create_app()


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
